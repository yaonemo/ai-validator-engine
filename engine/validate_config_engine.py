# -*- coding: utf-8 -*-
"""
通用配置驱动校验引擎（校验官引擎 v1.0）
=====================================
读取"校验配置JSON"（检查清单）并逐条执行校验。
引擎本身不含任何业务逻辑，业务规则全部来自配置文件。

用途：任何部门提供校验配置（可由AI根据部门标准文档生成），
即可用本引擎对数据进行稳定校验（对错100%可复现）。

用法:
    python3 validate_config_engine.py <输入Excel> <校验配置JSON> [输出报告路径]

规则类型（配置中的 type 字段）:
    required             必填（unless 条件满足时跳过）
    in_list              值必须在列表中（空值跳过）
    not_contains_chinese 值不能包含中文
    numeric              值必须是数字（allow_values 里的值除外）
    if_then              if 条件成立时 then 断言必须成立
    personal_name_warn   when 条件成立时，值疑似个人姓名 → warning
    org_name_check       when 条件成立时，值必须是组织/公司名
    not_contains_any     when 条件成立时，值不能包含任一关键词
    keyword_not          remark含关键词时，目标字段不能包含禁用关键词
    keyword_if           remark含关键词时，if 条件成立则报错
    keyword_and          remark含关键词 且 if 条件成立则报错

操作符（op）:
    eq / neq / in / not_in / contains / not_blank
"""

import json
import re
import sys
from collections import Counter
from typing import Dict, List, Tuple, Any, Optional

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl：请先安装依赖  python3 -m pip install -r requirements.txt")
    sys.exit(1)

# ---------------------------------------------------------------- 工具函数

def _clean(v: Any) -> str:
    """字段值统一转字符串并去空白"""
    if v is None:
        return ""
    return str(v).strip()


def _matches(actual: str, cond: Dict) -> bool:
    """判断字段值是否满足条件"""
    op = cond.get("op", "eq")
    target = cond.get("value")
    if op == "eq":
        return actual == target
    if op == "neq":
        return actual != target
    if op == "in":
        return actual in target if isinstance(target, list) else actual == target
    if op == "not_in":
        return actual not in target if isinstance(target, list) else actual != target
    if op == "contains":
        return target in actual
    if op == "not_blank":
        return actual != ""
    return False


def _fill_template(template: str, row_values: Dict[str, str], current_value: str = "") -> str:
    """替换消息模板中的 {value} 和 {field_key} 占位符"""
    msg = template.replace("{value}", current_value)
    for key, val in row_values.items():
        msg = msg.replace("{" + key + "}", val)
    return msg


def _is_personal_name(value: str, org_keywords: List[str]) -> bool:
    """判断是否为疑似个人姓名：纯汉字2-4字，且不含组织关键词"""
    if not value:
        return False
    if any(kw in value for kw in org_keywords):
        return False
    if value in ("无", "全体业主", "/"):
        return False
    return bool(re.match(r'^[\u4e00-\u9fff]{2,4}$', value))


# ---------------------------------------------------------------- 结构校验

def check_structure(ws, config: Dict) -> Tuple[List[str], Optional[Dict]]:
    """校验表头结构，返回 (问题列表, 列映射或None)"""
    standard = config["structure"]["标准表头"]
    aliases = config["structure"].get("表头别名", {})
    issues = []
    headers = [cell.value for cell in ws[1]]
    non_empty = [h for h in headers if h is not None]

    if len(non_empty) != len(standard):
        issues.append(f"列数不正确：标准{len(standard)}列，实际{len(non_empty)}列")

    for i, std in enumerate(standard):
        if i < len(headers) and headers[i] is not None:
            actual = str(headers[i]).strip()
            allowed = [std] + aliases.get(std, [])
            if actual not in allowed:
                issues.append(f"第{i+1}列字段名不匹配：标准为「{std}」，实际为「{actual}」")
        else:
            issues.append(f"缺少第{i+1}列：「{std}」")

    if issues:
        return issues, None

    # 构建列映射（精确匹配 + 模糊匹配）
    field_map = config["structure"]["字段映射"]
    col_map: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        h = str(header).strip()
        matched = None
        # 先精确匹配字段映射表
        for key, name in field_map.items():
            if h == name:
                matched = key
                break
        # 模糊匹配（按字段名长度降序，避免短名如"产权人"吃掉"产权人联系方式"）
        if not matched:
            for key, name in sorted(field_map.items(), key=lambda x: -len(x[1])):
                if name in h:
                    matched = key
                    break
        if matched and matched not in col_map:
            col_map[matched] = idx
    return [], col_map


# ---------------------------------------------------------------- 规则执行

def run_field_checks(rule, row_values: Dict[str, str], org_keywords: List[str]) -> List[Dict]:
    """执行一条字段/逻辑/场景校验规则，返回问题列表"""
    issues = []
    rtype = rule["type"]
    severity = rule.get("severity", "error")
    message = rule.get("message", "")
    field = rule.get("field")
    value = row_values.get(field, "") if field else ""

    # 条件触发型规则：when 不满足时跳过
    when = rule.get("when")
    if when and not _matches(row_values.get(when.get("field", ""), ""), when):
        return issues

    # unless：满足条件时跳过（用于 required 的可选场景）
    unless = rule.get("unless")
    if unless and _matches(row_values.get(unless.get("field", ""), ""), unless):
        return issues

    if rtype == "required":
        if value == "":
            issues.append({"field": rule.get("field_label", field), "severity": severity,
                           "message": _fill_template(message, row_values, value)})

    elif rtype == "in_list":
        if value != "" and value not in rule.get("list", []):
            issues.append({"field": rule.get("field_label", field), "severity": severity,
                           "message": _fill_template(message, row_values, value)})

    elif rtype == "not_contains_chinese":
        if value and re.search(r'[\u4e00-\u9fff]', value):
            issues.append({"field": rule.get("field_label", field), "severity": severity,
                           "message": _fill_template(message, row_values, value)})

    elif rtype == "numeric":
        allow = rule.get("allow_values", [])
        if value != "" and value not in allow:
            if not re.match(r'^[0-9]+\.?[0-9]*$', value):
                issues.append({"field": rule.get("field_label", field), "severity": severity,
                               "message": _fill_template(message, row_values, value)})

    elif rtype == "if_then":
        cond = rule.get("if", {})
        assert_cond = rule.get("then", {})
        if _matches(row_values.get(cond.get("field", ""), ""), cond):
            if not _matches(row_values.get(assert_cond.get("field", ""), ""), assert_cond):
                issues.append({"field": rule.get("field_label", "逻辑一致性"), "severity": severity,
                               "message": _fill_template(message, row_values, value)})

    elif rtype == "personal_name_warn":
        if value and _is_personal_name(value, org_keywords):
            issues.append({"field": rule.get("field_label", "逻辑一致性"), "severity": severity,
                           "message": _fill_template(message, row_values, value)})

    elif rtype == "org_name_check":
        if value and value != "无":
            if _is_personal_name(value, org_keywords) or value == "全体业主":
                issues.append({"field": rule.get("field_label", "逻辑一致性"), "severity": severity,
                               "message": _fill_template(message, row_values, value)})

    elif rtype == "not_contains_any":
        if value:
            for kw in rule.get("keywords", []):
                if kw in value:
                    issues.append({"field": rule.get("field_label", "逻辑一致性"), "severity": severity,
                                   "message": _fill_template(message, row_values, value)})
                    break

    elif rtype == "keyword_not":
        remark = row_values.get(rule.get("remark_field", "remark"), "")
        if rule.get("keyword") in remark:
            target = row_values.get(rule.get("target_field", ""), "")
            for kw in rule.get("forbid_keywords", []):
                if kw in target:
                    issues.append({"field": rule.get("field_label", "逻辑一致性"), "severity": severity,
                                   "message": _fill_template(message, row_values, value)})
                    break

    elif rtype == "keyword_if":
        remark = row_values.get(rule.get("remark_field", "remark"), "")
        if rule.get("keyword") in remark:
            cond = rule.get("if", {})
            if _matches(row_values.get(cond.get("field", ""), ""), cond):
                issues.append({"field": rule.get("field_label", "场景校验"), "severity": severity,
                               "message": _fill_template(message, row_values, value)})

    elif rtype == "keyword_and":
        remark = row_values.get(rule.get("remark_field", "remark"), "")
        cond = rule.get("if", {})
        if rule.get("keyword") in remark and _matches(row_values.get(cond.get("field", ""), ""), cond):
            issues.append({"field": rule.get("field_label", "逻辑一致性"), "severity": severity,
                           "message": _fill_template(message, row_values, value)})

    return issues


# ---------------------------------------------------------------- 报告生成

def merge_issues(issues: List[Dict]) -> List[Dict]:
    """合并相同 message 的问题，汇总行号范围"""
    groups: Dict[str, Dict] = {}
    order: List[str] = []
    for it in issues:
        key = (it["field"], it["severity"], it["message"])
        if key not in groups:
            groups[key] = {"field": it["field"], "severity": it["severity"],
                           "message": it["message"], "rows": []}
            order.append(key)
        groups[key]["rows"].append(it["row"])
    merged = []
    for key in order:
        g = groups[key]
        rows = sorted(set(g["rows"]))
        # 压缩连续行号
        ranges = []
        start = prev = rows[0]
        for r in rows[1:]:
            if r == prev + 1:
                prev = r
            else:
                ranges.append((start, prev))
                start = prev = r
        ranges.append((start, prev))
        row_desc = "、".join(f"{a}-{b}" if a != b else str(a) for a, b in ranges)
        g["row_desc"] = row_desc
        merged.append(g)
    return merged


def generate_report(project_name: str, total: int, valid: int, skipped: int,
                    merged: List[Dict]) -> str:
    """生成与业务校验器兼容的报告文本"""
    errors = [m for m in merged if m["severity"] == "error"]
    warnings = [m for m in merged if m["severity"] == "warning"]

    report = f"📋 **验证报告 - {project_name}**\n\n"
    report += f"总记录数: {total}，有效记录: {valid}，跳过行数: {skipped}\n"
    if total:
        report += f"通过率: {valid / total * 100:.1f}%\n"
    report += "\n"

    if not errors and not warnings:
        report += "✅ 全部通过，数据无误。"
        return report

    if errors:
        report += f"❌ **发现 {len(errors)} 个错误需要修改：**\n"
        for m in errors:
            report += f"• **第{m['row_desc']}行** {m['message']}\n"
        report += "\n"

    if warnings:
        report += f"⚠️ **发现 {len(warnings)} 个警告建议修改：**\n"
        for m in warnings:
            report += f"• **第{m['row_desc']}行** {m['message']}\n"
        report += "\n"

    if errors:
        report += "**有错误必须修正后重新提交，警告建议一并优化。**"
    else:
        report += "**数据基本可用，建议按警告内容优化。**"
    return report


# ---------------------------------------------------------------- 主流程

def validate_file(file_path: str, config_path: str, file_name: str = "") -> Tuple[str, List[Dict], bool]:
    """校验单个文件，返回 (报告, 问题列表, 是否全部通过)"""
    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    if not file_name:
        file_name = file_path.rsplit("/", 1)[-1]

    # 1. 结构校验
    struct_issues, col_map = check_structure(ws, config)
    if struct_issues:
        report = f"📋 **验证报告 - {file_name}**\n\n❌ **数据结构校验不通过：**\n\n"
        for issue in struct_issues:
            report += f"• {issue}\n"
        report += "\n**请使用标准模板，不要修改字段结构。**"
        return report, [], False

    # 2. 项目名提取（most_frequent）
    org_keywords = config.get("组织关键词", [])
    project_name = file_name
    name_field = config.get("项目名提取", {}).get("字段", "parking_name")
    name_counts: Counter = Counter()
    if name_field in col_map:
        nidx = col_map[name_field]
        for row in ws.iter_rows(min_row=config["structure"]["数据起始行"], values_only=True):
            if not any(row):
                continue
            if nidx < len(row) and row[nidx]:
                name_counts[str(row[nidx]).strip()] += 1
    if name_counts:
        project_name = name_counts.most_common(1)[0][0]

    # 3. 逐行校验
    issues: List[Dict] = []
    total_rows = 0
    valid_rows = 0
    skipped_rows = 0
    filter_cfg = config["structure"]["示例行过滤"]
    filter_field = filter_cfg.get("字段", "row_num")
    filter_kws = filter_cfg.get("包含关键词", ["示例"])

    field_checks = config.get("字段校验", [])
    logic_checks = config.get("逻辑校验", [])
    scenario_checks = config.get("场景校验", [])

    for row_idx, row in enumerate(ws.iter_rows(min_row=config["structure"]["数据起始行"], values_only=True),
                                  start=config["structure"]["数据起始行"]):
        if not any(row):
            continue
        # 示例行过滤
        a_val = _clean(row[0] if len(row) > 0 else "")
        if any(kw in a_val for kw in filter_kws):
            skipped_rows += 1
            continue

        total_rows += 1
        row_values: Dict[str, str] = {}
        for key, idx in col_map.items():
            row_values[key] = _clean(row[idx]) if idx < len(row) else ""

        row_issues = []
        for rule in field_checks + logic_checks + scenario_checks:
            row_issues.extend(run_field_checks(rule, row_values, org_keywords))

        # 附加行号到每个问题
        for it in row_issues:
            it["row"] = row_idx

        if row_issues:
            issues.extend(row_issues)
        else:
            valid_rows += 1

    # 4. 报告
    all_passed = all(it["severity"] != "error" for it in issues)
    merged = merge_issues(issues)
    report = generate_report(project_name, total_rows, valid_rows, skipped_rows, merged)
    return report, issues, all_passed


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python3 validate_config_engine.py <输入Excel> <校验配置JSON> [输出报告路径]")
        sys.exit(1)
    in_file = sys.argv[1]
    cfg_file = sys.argv[2]
    out_file = sys.argv[3] if len(sys.argv) > 3 else None

    report, issues, passed = validate_file(in_file, cfg_file)
    print(report)
    if out_file:
        with open(out_file, "w", encoding="utf-8") as f:
            f.write(report)
    print(f"\n[engine] passed={passed} issues={len(issues)}")
